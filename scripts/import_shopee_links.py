#!/usr/bin/env python3
"""
import_shopee_links.py — แปลง CSV "ลิงก์สินค้าหลายลิงก์" จาก Shopee Affiliate
                         แล้ว append rows เข้า picks.xlsx

โครงสร้าง folder แนะนำ:
    csv-affiliate-shoppe/products/
        ai_calculator/        ← ชื่อ folder = section
            5090.csv          ← ชื่อไฟล์ = ids
            5080.csv
        ai_calculator_cpu/
            r9_7950x.csv

Auto-mapping:
    section  = ชื่อ folder แม่ของไฟล์ CSV (override ด้วย --section)
    ids      = ชื่อไฟล์ (stem) (override ด้วย --ids)

Usage:
    # products root — ประมวลผลทุก section/ไฟล์ อัตโนมัติ
    python3 scripts/import_shopee_links.py csv-affiliate-shoppe/products/

    # folder เดียว — section = ชื่อ folder
    python3 scripts/import_shopee_links.py csv-affiliate-shoppe/products/ai_calculator/

    # single file — section = ชื่อ folder แม่, ids = ชื่อไฟล์
    python3 scripts/import_shopee_links.py csv-affiliate-shoppe/products/ai_calculator/5090.csv

    # override section / ids
    python3 scripts/import_shopee_links.py <path>  --section ev --ids "5090|5080"
    python3 scripts/import_shopee_links.py <path>  --dry-run

Options:
    --section      override section key (default: ชื่อ folder แม่)
    --source       Shopee | Lazada | Amazon | ...  (default: Shopee)
    --badge        ขายดี | แนะนำ | ราคาดี | ใหม่ | HOT  (optional)
    --ids          override ids ทุกไฟล์ (default: ชื่อไฟล์ stem)
    --hint         คำอธิบายสั้น ใส่ทุกแถว  (optional)
    --type         item | guide  (default: item)
    --limit        จำกัดจำนวนแถวสูงสุด (รวมทุกไฟล์)
    --feed-csv     path ของ Product Feed CSV ใหญ่ เพื่อดึงรูปสินค้า
    --use-product-link  ใช้ลิงก์สินค้าเต็มแทน affiliate short link
    --replace      แทนที่แถวใน section เดิมใน picks.xlsx (default: append)
    --dry-run      แสดงผลโดยไม่บันทึก
    --output       path ของ picks.xlsx ที่ต้องการเขียน
"""

import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print('❌ ต้องติดตั้ง openpyxl: pip3 install openpyxl --break-system-packages')
    sys.exit(1)

REPO_ROOT   = Path(__file__).parent.parent
DEFAULT_XLS = REPO_ROOT / 'data' / 'affiliate' / 'picks.xlsx'

FEED_COLS = (
    'image_link_4,cb_option,global_category1,stock,item_sold,is_preferred_shop,'
    'title,shopid,model_prices,global_category2,sale_price,like,shopee_verified_flag,'
    'global_catid2,is_item_welcome_package,is_official_shop,global_catid3,description,'
    'seller_penalty_score,global_catid1,image_link_3,additional_image_link,image_link_9,'
    'image_link_7,global_category3,global_item_attributes,condition,discount_percentage,'
    'model_names,image_link_6,holiday_mode_on,itemid,model_ids,has_lowest_price_guarantee,'
    'shop_sku_count,image_link_10,shop_rating,item_rating,seller_name,price,image_link_8,'
    'image_link,shop_name,global_brand,image_link_5,product_link,product_short link'
).split(',')

XLSX_HEADERS = [
    'section', 'type', 'ids', 'title',
    'price', 'original_price', 'link', 'image',
    'source', 'badge', 'shop_name', 'item_sold', 'row', 'hint',
]

# ─── ราคา helpers ────────────────────────────────────────────────────────────

_THAI_UNIT = {
    'พัน': 1_000,
    'หมื่น': 10_000,
    'แสน': 100_000,
    'ล้าน': 1_000_000,
}

def parse_price(raw: str) -> 'int | float | None':
    """แปลง '169.7พัน' → 169700, '฿1,697.30' → 1697, '9,990' → 9990"""
    s = str(raw).strip().replace(',', '').replace('฿', '').replace(' ', '')
    if not s:
        return None
    for suffix, mul in _THAI_UNIT.items():
        if s.endswith(suffix):
            try:
                v = float(s[:-len(suffix)]) * mul
                return int(round(v))
            except ValueError:
                return None
    try:
        v = float(s)
        return int(v) if v == int(v) else v
    except (ValueError, TypeError):
        return None

# ─── Input resolver ──────────────────────────────────────────────────────────

def resolve_inputs(raw: str, section_override: str) -> list[tuple[Path, str]]:
    """
    แปลง argument เป็นรายการ (csv_path, section) tuples

    กฎการหา section (ตามลำดับ priority):
      1. --section override
      2. ชื่อ folder แม่ของไฟล์ CSV (parent.name)

    กฎการ traverse:
      - single file     → [(file, section)]
      - folder มี CSV   → [(csv, section)] ทุกไฟล์ใน folder นั้น
      - folder มี subdir → recurse เข้า subdir แต่ละอัน (1 ชั้น)
    """
    p = Path(raw)
    if not p.is_absolute() and not p.exists():
        candidate = REPO_ROOT / p
        if candidate.exists():
            p = candidate

    if p.is_file():
        section = section_override or p.parent.name
        return [(p, section)]

    if p.is_dir():
        csv_files = sorted(p.glob('*.csv'))
        if csv_files:
            # folder ที่มี CSV โดยตรง → section = ชื่อ folder นี้
            section = section_override or p.name
            return [(f, section) for f in csv_files]

        # folder ที่มี subdir → แต่ละ subdir คือ section
        subdirs = sorted(d for d in p.iterdir() if d.is_dir() and not d.name.startswith('.'))
        results: list[tuple[Path, str]] = []
        for subdir in subdirs:
            section = section_override or subdir.name
            for f in sorted(subdir.glob('*.csv')):
                results.append((f, section))
        if not results:
            print(f'❌ ไม่พบ .csv หรือ subfolder ใน: {p}')
            sys.exit(1)
        return results

    print(f'❌ ไม่พบไฟล์หรือ folder: {p}')
    sys.exit(1)

# ─── CSV reader ───────────────────────────────────────────────────────────────

def read_shopee_links_csv(path: Path) -> list[dict]:
    """อ่าน CSV 'ลิงก์สินค้าหลายลิงก์' จาก Shopee Affiliate Portal"""
    rows = []
    with open(path, encoding='utf-8-sig', errors='replace', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not any(v.strip() for v in row.values()):
                continue
            rows.append({k.strip(): v.strip() for k, v in row.items()})
    return rows

# ─── Product Feed image lookup ───────────────────────────────────────────────

def build_image_index(feed_csv: Path) -> dict:
    """สร้าง dict {itemid: image_url} จาก Product Feed CSV ใหญ่"""
    index: dict = {}
    print(f'🔍 กำลังสร้าง image index จาก {feed_csv.name}...')
    try:
        with open(feed_csv, encoding='utf-8-sig', errors='replace') as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                if len(row) < 42:
                    continue
                d = dict(zip(FEED_COLS, row))
                itemid = d.get('itemid', '').strip()
                img    = d.get('image_link', '').strip()
                if itemid and img:
                    index[itemid] = img
    except FileNotFoundError:
        print(f'⚠ ไม่พบ feed CSV: {feed_csv}')
    print(f'   พบ {len(index):,} รายการใน feed')
    return index

# ─── Row builder ─────────────────────────────────────────────────────────────

def build_xlsx_row(
    src: dict,
    *,
    section: str,
    source: str,
    badge: str,
    ids: str,
    hint: str,
    row_type: str,
    use_product_link: bool,
    image_index: dict,
) -> list:
    """แปลง 1 แถวจาก Shopee CSV → 1 row สำหรับ picks.xlsx"""
    itemid    = src.get('รหัสสินค้า', '').strip()
    title     = src.get('ชื่อสินค้า', '').strip()
    price     = parse_price(src.get('ราคา', ''))
    shop_name = src.get('ชื่อร้านค้า', '').strip()

    try:
        item_sold = int(str(src.get('ขาย', '0')).strip()) or None
    except (ValueError, TypeError):
        item_sold = None

    if use_product_link:
        link = src.get('ลิงก์สินค้า', '').strip()
    else:
        link = src.get('ลิงก์ข้อเสนอ', '').strip() or src.get('ลิงก์สินค้า', '').strip()

    # รูปสินค้า: priority 1=คอลัมน์ในCSV, 2=feed-csv lookup
    image = (
        src.get('รูปสินค้า', '').strip()
        or image_index.get(itemid, '')
    )

    return [
        section,
        row_type,
        ids,
        title,
        price,
        None,             # original_price
        link,
        image or None,
        source,
        badge or None,
        shop_name or None,
        item_sold,
        None,             # row (guide only)
        hint or None,
    ]

# ─── xlsx append/replace ──────────────────────────────────────────────────────

def append_to_xlsx(
    xlsx_path: Path,
    new_rows: list[list],
    sections_to_replace: set[str],
    replace: bool,
) -> tuple:
    """
    append (หรือ replace) rows ใน picks.xlsx
    sections_to_replace — set ของ section ที่ต้องลบแถวเดิมก่อน (ใช้เมื่อ replace=True)
    Returns: (rows_added, rows_removed)
    """
    wb = load_workbook(xlsx_path)
    ws = wb['Picks'] if 'Picks' in wb.sheetnames else wb.active

    header_row = [str(c.value).strip() if c.value else '' for c in ws[1]]

    removed = 0
    if replace and sections_to_replace:
        sec_col = header_row.index('section') + 1 if 'section' in header_row else 1
        rows_to_delete = [
            r for r in range(2, ws.max_row + 1)
            if str(ws.cell(row=r, column=sec_col).value or '').strip() in sections_to_replace
        ]
        for r in reversed(rows_to_delete):
            ws.delete_rows(r)
        removed = len(rows_to_delete)

    for data_row in new_rows:
        ws.append(data_row)

    wb.save(xlsx_path)
    return len(new_rows), removed

# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='แปลง Shopee Affiliate "ลิงก์สินค้าหลายลิงก์" CSV → picks.xlsx\n'
                    'รับได้ทั้ง single file และ folder (ประมวลผลทุก *.csv ในคราวเดียว)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        'input',
        help='path ของ CSV หรือ folder ที่มี *.csv (เช่น csv-affiliate-shoppe/products/)',
    )
    parser.add_argument('--section', default='', metavar='KEY',
                        help='override section key (default: ชื่อ folder แม่ของ CSV)')
    parser.add_argument('--source', default='Shopee', metavar='PLATFORM',
                        help='ชื่อ platform: Shopee|Lazada|Amazon|JD|อื่นๆ  (default: Shopee)')
    parser.add_argument('--badge', default='', metavar='LABEL',
                        help='badge label: ขายดี|แนะนำ|ราคาดี|ใหม่|HOT  (optional)')
    parser.add_argument('--ids', default='', metavar='ID1|ID2',
                        help='override ids ทุกไฟล์ (default: ใช้ชื่อไฟล์ stem เช่น 5090.csv → "5090")')
    parser.add_argument('--hint', default='', metavar='TEXT',
                        help='คำอธิบายสั้นที่ใส่ทุกแถว  (optional)')
    parser.add_argument('--type', default='item', choices=['item', 'guide'], dest='row_type',
                        help='ประเภท: item (default) | guide')
    parser.add_argument('--limit', type=int, default=None, metavar='N',
                        help='จำกัดจำนวนแถวสูงสุดรวมทุกไฟล์')
    parser.add_argument('--feed-csv', default='', metavar='PATH',
                        help='path ของ Product Feed CSV เพื่อดึงรูปสินค้า (optional)')
    parser.add_argument('--use-product-link', action='store_true',
                        help='ใช้ลิงก์สินค้าเต็ม แทน affiliate short link')
    parser.add_argument('--replace', action='store_true',
                        help='แทนที่แถวของ section นั้นใน picks.xlsx (default: append)')
    parser.add_argument('--dry-run', action='store_true',
                        help='แสดงผลโดยไม่บันทึก')
    parser.add_argument('--output', default=str(DEFAULT_XLS), metavar='PATH',
                        help='path ของ picks.xlsx (default: data/affiliate/picks.xlsx)')

    args = parser.parse_args()

    # ─── ตรวจ output xlsx ──────────────────────────────────────────────────────
    output_path = Path(args.output)
    if not output_path.is_absolute():
        candidate = REPO_ROOT / output_path
        if candidate.exists():
            output_path = candidate
    if not output_path.exists():
        print(f'❌ ไม่พบ picks.xlsx: {output_path}')
        print(f'   สร้าง template ใหม่: python3 scripts/create_picks_xlsx.py')
        sys.exit(1)

    # ─── resolve input files + sections ───────────────────────────────────────
    input_files = resolve_inputs(args.input, args.section)  # list[(Path, section)]

    unique_sections = list(dict.fromkeys(sec for _, sec in input_files))

    if len(input_files) == 1:
        fp, sec = input_files[0]
        print(f'📄 ไฟล์: {fp.name}  (section={sec})')
    else:
        print(f'📁 พบ {len(input_files)} ไฟล์ ({len(unique_sections)} section):')
        for fp, sec in input_files:
            print(f'   • {sec}/{fp.name}')

    # ─── image index (สร้างครั้งเดียว ใช้กับทุกไฟล์) ──────────────────────────
    image_index: dict = {}
    if args.feed_csv:
        feed_path = Path(args.feed_csv)
        if not feed_path.is_absolute() and not feed_path.exists():
            feed_path = REPO_ROOT / feed_path
        image_index = build_image_index(feed_path)

    # ─── อ่านและ build rows จากทุกไฟล์ ───────────────────────────────────────
    all_new_rows: list[list] = []
    total_skipped  = 0
    per_file_stats = []

    for file_path, effective_section in input_files:
        src_rows = read_shopee_links_csv(file_path)
        file_ok  = 0
        file_skip = 0

        # ids: ใช้ --ids ที่ระบุมา หรือ fallback เป็น stem ของชื่อไฟล์
        effective_ids = args.ids if args.ids else file_path.stem

        for i, src in enumerate(src_rows, start=1):
            title = src.get('ชื่อสินค้า', '').strip()
            price = parse_price(src.get('ราคา', ''))
            link  = (
                src.get('ลิงก์ข้อเสนอ', '').strip()
                if not args.use_product_link
                else src.get('ลิงก์สินค้า', '').strip()
            )

            if not title or not link or price is None:
                file_skip += 1
                total_skipped += 1
                continue

            row = build_xlsx_row(
                src,
                section=effective_section,
                source=args.source,
                badge=args.badge,
                ids=effective_ids,
                hint=args.hint,
                row_type=args.row_type,
                use_product_link=args.use_product_link,
                image_index=image_index,
            )
            all_new_rows.append(row)
            file_ok += 1

        per_file_stats.append((file_path.name, effective_section, effective_ids, len(src_rows), file_ok, file_skip))

    # ─── apply global limit ────────────────────────────────────────────────────
    if args.limit and len(all_new_rows) > args.limit:
        all_new_rows = all_new_rows[:args.limit]
        print(f'   → จำกัดที่ {args.limit} แถว (รวมทุกไฟล์)')

    # ─── สรุปก่อน import ──────────────────────────────────────────────────────
    print(f'\n📋 สรุปก่อน import:')
    print(f'   section  : {"--section override: " + args.section if args.section else "auto (จากชื่อ folder)"}')
    print(f'   ids mode : {"--ids override: " + args.ids if args.ids else "auto (จากชื่อไฟล์)"}')
    print(f'   source   : {args.source}')
    print(f'   badge    : {args.badge or "(ว่าง)"}')
    print(f'   mode     : {"replace" if args.replace else "append"}')
    print(f'   ไฟล์ทั้งหมด : {len(input_files)} ไฟล์ ({len(unique_sections)} section)')

    if len(input_files) > 1:
        for name, eff_sec, eff_ids, total, ok, skip in per_file_stats:
            print(f'     • {eff_sec}/{name:<35} ids={eff_ids:<20} {ok} แถว ({skip} ข้าม)')
    else:
        name, eff_sec, eff_ids, total, ok, skip = per_file_stats[0]
        print(f'   section  : {eff_sec}')
        print(f'   ids      : {eff_ids}')

    print(f'   แถวที่ valid  : {len(all_new_rows)}')
    print(f'   แถวที่ข้าม   : {total_skipped}')
    print(f'   มีรูปสินค้า  : {sum(1 for r in all_new_rows if r[7])}')

    if args.dry_run:
        print(f'\n[dry-run] ตัวอย่าง 5 แถวแรก:')
        for r in all_new_rows[:5]:
            d = dict(zip(XLSX_HEADERS, r))
            print(f'  {d["title"][:55]:<55} ฿{str(d["price"]):>8}  {d["link"][:40]}')
        print('\n[dry-run] ไม่บันทึก — ลบ --dry-run เพื่อ import จริง')
        return

    if not all_new_rows:
        print('\n⚠ ไม่มีแถวที่ valid — ไม่บันทึก')
        return

    # ─── เขียนลง xlsx ─────────────────────────────────────────────────────────
    added, removed = append_to_xlsx(
        output_path,
        all_new_rows,
        sections_to_replace=set(unique_sections),
        replace=args.replace,
    )

    print(f'\n✅ บันทึกแล้ว → {output_path.relative_to(REPO_ROOT)}')
    if removed:
        print(f'   ลบแถวเดิม  : {removed} แถว')
    print(f'   เพิ่มแถวใหม่: {added} แถว')
    print(f'\n👉 ขั้นตอนถัดไป: รัน import_picks.py เพื่อ sync ไปยัง JSON')
    if len(unique_sections) == 1:
        print(f'   python3 scripts/import_picks.py --section {unique_sections[0]}')
    else:
        print(f'   python3 scripts/import_picks.py  # (sync ทุก section)')

if __name__ == '__main__':
    main()
