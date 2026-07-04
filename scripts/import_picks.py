#!/usr/bin/env python3
"""
import_picks.py — นำเข้าข้อมูล affiliate จาก Excel/CSV → manual-picks.json

Excel/CSV เป็น Source of Truth:
  - รันสคริปต์ทุกครั้งที่แก้ไข picks.xlsx (หรือ picks.csv)
  - ทุก section ที่ปรากฏใน file จะถูก replace ใน JSON (section อื่น → ไม่ถูกแตะ)
  - รองรับ .xlsx (แนะนำ) และ .csv

Usage:
    python3 scripts/import_picks.py                     # ใช้ picks.xlsx (default)
    python3 scripts/import_picks.py picks.csv           # ใช้ CSV แทน
    python3 scripts/import_picks.py --dry-run           # preview ไม่บันทึก
    python3 scripts/import_picks.py --section mac_llm   # import แค่ section เดียว

Excel Columns (sheet ชื่อ "Picks"):
    section         key ใน manual-picks.json (ai_calculator, mac_llm, ...)
    type            item | guide
    ids             id คั่นด้วย | เช่น r9_7950x|r9_7900x  (items + filterId)
    title           ชื่อสินค้า (items) / ชื่อสินค้าใน guide column (guides)
    price           ราคา (ตัวเลข)
    original_price  ราคาก่อนลด (optional, items)
    link            affiliate URL
    image           URL รูปสินค้า (optional, items)
    source          Shopee | Lazada | ฯลฯ (optional, items)
    badge           ป้าย: ขายดี | แนะนำ | ราคาดี (optional, items)
    row             row index 0-based (guides)
    hint            คำอธิบายสั้น (optional, guides)
"""

import csv
import json
import sys
import argparse
from pathlib import Path

REPO_ROOT   = Path(__file__).parent.parent
PICKS_JSON  = REPO_ROOT / 'data' / 'affiliate' / 'manual-picks.json'
DEFAULT_XLS = REPO_ROOT / 'data' / 'affiliate' / 'picks.xlsx'
DEFAULT_CSV = REPO_ROOT / 'data' / 'affiliate' / 'picks.csv'

EXPECTED_HEADERS = [
    'section', 'type', 'ids', 'title',
    'price', 'original_price', 'link', 'image',
    'source', 'badge', 'row', 'hint',
]

# ─── helpers ──────────────────────────────────────────────────────────────────

def _str(v) -> str:
    if v is None:
        return ''
    return str(v).strip()

def _price(raw) -> int | float | None:
    s = _str(raw).replace(',', '')
    if not s:
        return None
    try:
        v = float(s)
        return int(v) if v == int(v) else v
    except (ValueError, TypeError):
        return None

def _ids(raw: str) -> list:
    return [i.strip() for i in _str(raw).split('|') if i.strip()]

def _build_item(row: dict) -> dict | None:
    title = _str(row.get('title'))
    link  = _str(row.get('link'))
    price = _price(row.get('price'))
    if not title or not link or price is None:
        return None

    item: dict = {'title': title, 'price': price, 'link': link}

    op = _price(row.get('original_price'))
    if op is not None and op > price:
        item['original_price'] = op

    for field in ('image', 'source', 'badge'):
        v = _str(row.get(field))
        if v:
            item[field] = v

    ids = _ids(_str(row.get('ids', '')))
    if ids:
        item['ids'] = ids

    return item

def _build_guide(row: dict) -> dict | None:
    name  = _str(row.get('title'))
    link  = _str(row.get('link'))
    price = _price(row.get('price'))
    try:
        r = int(float(_str(row.get('row', ''))))
    except (ValueError, TypeError):
        return None
    if not name or not link:
        return None

    guide: dict = {'row': r, 'name': name, 'link': link}
    if price is not None:
        guide['price'] = price
    hint = _str(row.get('hint'))
    if hint:
        guide['hint'] = hint
    return guide

# ─── readers ──────────────────────────────────────────────────────────────────

def read_xlsx(path: Path) -> list[dict]:
    """อ่าน sheet 'Picks' จาก Excel → list of row dicts"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('❌ ต้องติดตั้ง openpyxl: pip3 install openpyxl --break-system-packages')
        sys.exit(1)

    wb = load_workbook(path, data_only=True)
    # ใช้ sheet ชื่อ 'Picks' ก่อน, fallback sheet แรก
    ws = wb['Picks'] if 'Picks' in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h).strip() if h else '' for h in rows[0]]

    # ตรวจ header ครบไหม
    missing = [h for h in EXPECTED_HEADERS if h not in header]
    if missing:
        print(f'⚠ Excel columns ที่หายไป: {missing}')
        print(f'  header ที่พบ: {header}')

    result = []
    for data_row in rows[1:]:
        row_dict = {header[i]: data_row[i] for i in range(min(len(header), len(data_row)))}
        # ข้ามแถวว่าง
        if not any(_str(v) for v in row_dict.values()):
            continue
        result.append(row_dict)

    return result

def read_csv(path: Path) -> list[dict]:
    """อ่าน CSV → list of row dicts (ข้าม # comment lines)"""
    result = []
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(
            (line for line in f if not line.lstrip().startswith('#')),
        )
        for row in reader:
            if not any(_str(v) for v in row.values()):
                continue
            result.append(row)
    return result

# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='นำเข้า affiliate picks จาก Excel/CSV → manual-picks.json'
    )
    parser.add_argument('input_file', nargs='?',
                        help='path ไปยัง .xlsx หรือ .csv (default: picks.xlsx)')
    parser.add_argument('--dry-run', action='store_true',
                        help='แสดงผลลัพธ์โดยไม่บันทึก JSON')
    parser.add_argument('--section', metavar='KEY',
                        help='import เฉพาะ section นี้เท่านั้น')
    args = parser.parse_args()

    # เลือกไฟล์ input
    if args.input_file:
        input_path = Path(args.input_file)
    elif DEFAULT_XLS.exists():
        input_path = DEFAULT_XLS
    elif DEFAULT_CSV.exists():
        input_path = DEFAULT_CSV
    else:
        print(f'❌ ไม่พบ picks.xlsx หรือ picks.csv ใน data/affiliate/')
        sys.exit(1)

    if not input_path.exists():
        print(f'❌ ไม่พบไฟล์: {input_path}')
        sys.exit(1)

    # อ่านข้อมูล
    suffix = input_path.suffix.lower()
    if suffix == '.xlsx':
        raw_rows = read_xlsx(input_path)
        print(f'📂 อ่าน {input_path.name} ({len(raw_rows)} แถว)')
    elif suffix == '.csv':
        raw_rows = read_csv(input_path)
        print(f'📂 อ่าน {input_path.name} ({len(raw_rows)} แถว)')
    else:
        print(f'❌ รองรับเฉพาะ .xlsx หรือ .csv เท่านั้น')
        sys.exit(1)

    # โหลด JSON เดิม
    with open(PICKS_JSON, 'r', encoding='utf-8') as f:
        picks = json.load(f)

    # แยก items / guide ตาม section
    collected_items: dict[str, list] = {}
    collected_guide: dict[str, list] = {}
    errors = []

    for lineno, row in enumerate(raw_rows, start=2):
        section = _str(row.get('section'))
        rtype   = _str(row.get('type', 'item')).lower()

        if not section:
            continue
        if args.section and section != args.section:
            continue

        if rtype == 'guide':
            obj = _build_guide(row)
            if obj:
                collected_guide.setdefault(section, []).append(obj)
            else:
                errors.append(
                    f'  ⚠ แถว {lineno} [guide/{section}]: '
                    f'ข้อมูลไม่ครบ (title/link/row required)'
                )
        else:
            obj = _build_item(row)
            if obj:
                collected_items.setdefault(section, []).append(obj)
            else:
                errors.append(
                    f'  ⚠ แถว {lineno} [item/{section}]: '
                    f'ข้อมูลไม่ครบ (title/price/link required)'
                )

    if errors:
        print('\n'.join(errors))

    if not collected_items and not collected_guide:
        print('❌ ไม่พบข้อมูลที่ valid')
        sys.exit(1)

    # อัปเดต picks dict
    total = 0
    for section, items in collected_items.items():
        if section not in picks:
            picks[section] = {'label': f'{section}', 'items': []}
        picks[section]['items'] = items
        print(f'  ✅ {section}.items  → {len(items):>3} รายการ')
        total += len(items)

    for section, guide in collected_guide.items():
        if section not in picks:
            picks[section] = {'label': f'{section}', 'items': [], 'guide': []}
        guide.sort(key=lambda g: g['row'])
        picks[section]['guide'] = guide
        print(f'  ✅ {section}.guide  → {len(guide):>3} แถว')
        total += len(guide)

    if args.dry_run:
        print(f'\n[dry-run] จะ import {total} รายการ — ไม่บันทึก')
        print('─' * 60)
        preview = {k: v for k, v in picks.items() if not k.startswith('_')}
        print(json.dumps(preview, ensure_ascii=False, indent=2)[:3000])
        return

    with open(PICKS_JSON, 'w', encoding='utf-8') as f:
        json.dump(picks, f, ensure_ascii=False, indent=2)
        f.write('\n')

    print(f'\n✅ บันทึกแล้ว → {PICKS_JSON.relative_to(REPO_ROOT)} ({total} รายการ)')

if __name__ == '__main__':
    main()
