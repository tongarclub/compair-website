#!/usr/bin/env python3
"""
create_picks_xlsx.py — สร้าง picks.xlsx template พร้อม formatting + data validation

Usage:
    python3 scripts/create_picks_xlsx.py
    python3 scripts/create_picks_xlsx.py --output path/to/file.xlsx
"""

import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo

except ImportError:
    print('❌ ต้องติดตั้ง openpyxl ก่อน: pip3 install openpyxl --break-system-packages')
    raise

REPO_ROOT   = Path(__file__).parent.parent
DEFAULT_OUT = REPO_ROOT / 'data' / 'affiliate' / 'picks.xlsx'

# ─── สีและ Style ────────────────────────────────────────────────────────────────

HEADER_BG   = 'FF1F2937'   # dark slate
HEADER_FG   = 'FFFFFFFF'

SECTION_COLORS = {
    'ai_calculator':     'FFE8F4FD',   # ฟ้าอ่อน
    'ai_calculator_cpu': 'FFFDE8E8',   # ชมพูอ่อน
    'mac_llm':           'FFE8FDE8',   # เขียวอ่อน
    'solar':             'FFFFF3E0',   # เหลืองอ่อน
    'ev':                'FFF3E0FF',   # ม่วงอ่อน
    'gold':              'FFFFF8E1',   # ทอง
    'image_gen':         'FFE0F7FA',   # ฟ้าน้ำทะเล
}

GUIDE_STRIPE = 'FFF5F5F5'  # สำหรับ type=guide

# ─── ข้อมูลตัวอย่าง (example rows) ─────────────────────────────────────────────

EXAMPLE_ROWS = [
    # section,            type,   ids,                                     title,                                      price,  orig,    link,                                       image, source, badge, row, hint
    ('ai_calculator',     'item', '',                                      'ASUS DUAL GeForce RTX 5060 8GB GDDR7',    14990,  16500,   'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    'Shopee','ขายดี','', ''),
    ('ai_calculator',     'item', '',                                      'MSI GeForce RTX 5060 VENTUS 2X 8G',       13990,  '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    'Shopee','',     '', ''),
    ('ai_calculator_cpu', 'item', 'r9_7950x|r9_7900x',                    'AMD Ryzen 9 7950X Box',                   18990,  '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    'Shopee','แนะนำ','', ''),
    ('ai_calculator_cpu', 'item', 'r7_7700x',                             'AMD Ryzen 7 7700X Box',                   11990,  '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    'Shopee','',     '', ''),
    ('ai_calculator_cpu', 'item', 'i9_14900k|i7_14700k',                  'Intel Core i9-14900K Box',                17500,  '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    'Shopee','ราคาดี','', ''),
    ('ai_calculator_cpu', 'item', 'r9_7950x|r9_7900x|r7_7700x|r5_7600x', 'ASUS PRIME X670-P WiFi Motherboard AM5',  8990,   '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    'Shopee','',     '', ''),
    ('mac_llm',           'item', '',                                      'Apple MacBook Air M4 13" 16GB',            39900,  '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    'Shopee','แนะนำ','', ''),
    ('mac_llm',           'guide','',                                      'เคส MacBook Air M4',                       590,    '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    '',     '',      0,  'ป้องกันรอย'),
    ('mac_llm',           'guide','',                                      'เคส MacBook Air M4 32GB',                  620,    '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    '',     '',      1,  'ป้องกันรอย'),
    ('mac_llm',           'guide','',                                      'เคส MacBook Pro M4 Pro',                   650,    '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    '',     '',      2,  'กันกระแทก'),
    ('mac_llm',           'guide','',                                      'เคส MacBook Pro M4 Pro 48GB',              650,    '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    '',     '',      3,  ''),
    ('mac_llm',           'guide','',                                      'เคส MacBook Pro M4 Max',                   680,    '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    '',     '',      4,  ''),
    ('mac_llm',           'guide','',                                      'เคส MacBook Pro M4 Max 128GB',             680,    '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    '',     '',      5,  ''),
    ('mac_llm',           'guide','',                                      'Stand + Hub Mac Studio',                   1290,   '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    '',     '',      6,  'ตั้งโต๊ะ + USB Hub'),
    ('solar',             'item', '',                                      'แผงโซล่าเซลล์ 550W Mono',                 4990,   '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    'Shopee','ขายดี','', ''),
    ('ev',                'item', '',                                      'EV Charger 7kW Wallbox Type 2',            12900,  '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    'Shopee','แนะนำ','', ''),
    ('gold',              'item', '',                                      'ทองคำแท่ง 1 บาท ออโรร่า',                 52000,  '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    'Shopee','',     '', ''),
    ('image_gen',         'item', '',                                      'ASUS DUAL GeForce RTX 5060 8GB',          14990,  '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    'Shopee','',     '', ''),
    ('image_gen',         'guide','',                                      'MSI GeForce RTX 5060 VENTUS 2X 8G',       13990,  '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    '',     '',      0,  'VRAM 8GB เริ่มต้น SD/FLUX Schnell'),
    ('image_gen',         'guide','',                                      'ASUS DUAL GeForce RTX 5060 Ti 16G',       19990,  '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    '',     '',      1,  'VRAM 16GB ครบ FLUX Dev FP8'),
    ('image_gen',         'guide','',                                      'MSI GeForce RTX 5070 Ti Gaming X Slim',   35990,  '',      'https://s.shopee.co.th/ใส่ลิงก์จริง',    '',    '',     '',      2,  'VRAM 16GB Pro batch gen'),
]

HEADERS = [
    'section', 'type', 'ids', 'title',
    'price', 'original_price', 'link', 'image',
    'source', 'badge', 'row', 'hint',
]

COL_WIDTHS = {
    'section':        18,
    'type':           8,
    'ids':            36,
    'title':          45,
    'price':          10,
    'original_price': 14,
    'link':           45,
    'image':          45,
    'source':         10,
    'badge':          10,
    'row':            5,
    'hint':           30,
}

# ─── Ref sheet ──────────────────────────────────────────────────────────────────

SECTION_NOTES = [
    ('ai_calculator',     'item',          'AI Calculator — GPU strip'),
    ('ai_calculator_cpu', 'item + ids',    'AI Calculator — CPU strip (กรองด้วย ids)'),
    ('mac_llm',           'item / guide',  'Mac LLM Calculator — strip + guide (row 0–6)'),
    ('solar',             'item',          'Solar Calculator — Solar strip'),
    ('ev',                'item',          'EV Calculator — EV Charger strip'),
    ('gold',              'item',          'Gold Calculator — ทองคำ strip'),
    ('image_gen',         'item / guide',  'Image Gen Calculator — strip + guide (row 0–2)'),
]

CPU_IDS = [
    ('r9_7950x',  'AMD Ryzen 9 7950X'),
    ('r9_7900x',  'AMD Ryzen 9 7900X'),
    ('r7_7700x',  'AMD Ryzen 7 7700X'),
    ('r5_7600x',  'AMD Ryzen 5 7600X'),
    ('r9_5900x',  'AMD Ryzen 9 5900X'),
    ('r7_5800x',  'AMD Ryzen 7 5800X'),
    ('r5_5600x',  'AMD Ryzen 5 5600X'),
    ('i9_14900k', 'Intel Core i9-14900K'),
    ('i7_14700k', 'Intel Core i7-14700K'),
    ('i5_14600k', 'Intel Core i5-14600K'),
    ('i9_13900k', 'Intel Core i9-13900K'),
    ('i7_13700k', 'Intel Core i7-13700K'),
    ('i5_13600k', 'Intel Core i5-13600K'),
    ('i9_12900k', 'Intel Core i9-12900K'),
    ('i7_12700k', 'Intel Core i7-12700K'),
]

MAC_ROWS = [
    (0, 'MacBook Air M4 16 GB'),
    (1, 'MacBook Air M4 32 GB'),
    (2, 'MacBook Pro M4 Pro 24 GB'),
    (3, 'MacBook Pro M4 Pro 48 GB'),
    (4, 'MacBook Pro M4 Max 36 GB'),
    (5, 'MacBook Pro M4 Max 128 GB'),
    (6, 'Mac Studio'),
]

GPU_ROWS = [
    (0, 'RTX 5060 8 GB — เริ่มต้น'),
    (1, 'RTX 5060 Ti 16 GB — จริงจัง'),
    (2, 'RTX 5070 Ti / 5080 — Pro'),
]

# ─── helpers ────────────────────────────────────────────────────────────────────

def hfill(hex_color: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_color)

def thin_border() -> Border:
    s = Side(style='thin', color='FFD1D5DB')
    return Border(left=s, right=s, top=s, bottom=s)

def header_font():
    return Font(name='Calibri', bold=True, color=HEADER_FG, size=10)

def body_font(bold=False):
    return Font(name='Calibri', bold=bold, size=10)

# ─── build workbook ─────────────────────────────────────────────────────────────

def build_workbook() -> Workbook:
    wb = Workbook()

    # ── sheet 1: Picks ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Picks'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'

    # Header row
    for col_idx, col_name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font    = header_font()
        cell.fill    = hfill(HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        cell.border  = thin_border()

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, row_data in enumerate(EXAMPLE_ROWS, start=2):
        section = row_data[0]
        rtype   = row_data[1]

        if rtype == 'guide':
            bg = GUIDE_STRIPE
        else:
            bg = SECTION_COLORS.get(section, 'FFFFFFFF')

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value != '' else None)
            cell.fill    = hfill(bg)
            cell.border  = thin_border()
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            # price columns → number format
            if col_idx in (5, 6, 11):
                cell.number_format = '#,##0'
            # link columns → เขียวอ่อน ถ้ามีค่า
            if col_idx in (7, 8) and value:
                cell.font = Font(name='Calibri', size=10, color='FF0969DA', underline='single')
            else:
                cell.font = body_font()

        ws.row_dimensions[row_idx].height = 18

    # Column widths
    for col_idx, col_name in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_name]

    # Data Validation — type
    dv_type = DataValidation(
        type='list',
        formula1='"item,guide"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle='ค่าไม่ถูกต้อง',
        error='ใส่ได้เฉพาะ item หรือ guide เท่านั้น',
    )
    dv_type.sqref = f'B2:B2000'
    ws.add_data_validation(dv_type)

    # Data Validation — source
    dv_source = DataValidation(
        type='list',
        formula1='"Shopee,Lazada,JD,Official,อื่นๆ"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_source.sqref = 'I2:I2000'
    ws.add_data_validation(dv_source)

    # Data Validation — badge
    dv_badge = DataValidation(
        type='list',
        formula1='"ขายดี,แนะนำ,ราคาดี,ใหม่,HOT"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_badge.sqref = 'J2:J2000'
    ws.add_data_validation(dv_badge)

    # Data Validation — row (0–9)
    dv_row = DataValidation(
        type='whole',
        operator='between',
        formula1='0',
        formula2='9',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle='row ไม่ถูกต้อง',
        error='row ต้องเป็นตัวเลข 0–9',
    )
    dv_row.sqref = 'K2:K2000'
    ws.add_data_validation(dv_row)

    # ── sheet 2: Ref ────────────────────────────────────────────────────────────
    ref = wb.create_sheet('Ref')
    ref.sheet_view.showGridLines = False

    def ref_header(text, row, col, colspan=1):
        cell = ref.cell(row=row, column=col, value=text)
        cell.font   = Font(name='Calibri', bold=True, size=10, color=HEADER_FG)
        cell.fill   = hfill(HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border()

    def ref_cell(text, row, col, bold=False, bg='FFFFFFFF'):
        cell = ref.cell(row=row, column=col, value=text)
        cell.font   = Font(name='Calibri', bold=bold, size=10)
        cell.fill   = hfill(bg)
        cell.alignment = Alignment(vertical='center', wrap_text=False)
        cell.border = thin_border()

    # Section table
    ref_header('section (key)', 1, 1)
    ref_header('type', 1, 2)
    ref_header('คำอธิบาย', 1, 3)
    ref.column_dimensions['A'].width = 22
    ref.column_dimensions['B'].width = 16
    ref.column_dimensions['C'].width = 44

    for i, (sec, typ, desc) in enumerate(SECTION_NOTES, start=2):
        bg = SECTION_COLORS.get(sec, 'FFFFFFFF')
        ref_cell(sec,  i, 1, bold=True, bg=bg)
        ref_cell(typ,  i, 2, bg=bg)
        ref_cell(desc, i, 3, bg=bg)
        ref.row_dimensions[i].height = 18

    # CPU IDs table (col 5–6)
    ref_header('ids value', 1, 5)
    ref_header('CPU', 1, 6)
    ref.column_dimensions['E'].width = 14
    ref.column_dimensions['F'].width = 28

    for i, (cpu_id, cpu_name) in enumerate(CPU_IDS, start=2):
        bg = 'FFFDE8E8' if i % 2 == 0 else 'FFFFFFFF'
        ref_cell(cpu_id,   i, 5, bold=True, bg=bg)
        ref_cell(cpu_name, i, 6, bg=bg)
        ref.row_dimensions[i].height = 18

    # mac_llm guide rows (col 8–9)
    ref_header('row (mac_llm)', 1, 8)
    ref_header('Mac tier', 1, 9)
    ref.column_dimensions['H'].width = 16
    ref.column_dimensions['I'].width = 28

    for i, (r, label) in enumerate(MAC_ROWS, start=2):
        bg = 'FFE8FDE8' if i % 2 == 0 else 'FFFFFFFF'
        ref_cell(r,     i, 8, bold=True, bg=bg)
        ref_cell(label, i, 9, bg=bg)

    # image_gen guide rows (col 11–12)
    ref_header('row (image_gen)', 1, 11)
    ref_header('GPU tier', 1, 12)
    ref.column_dimensions['K'].width = 18
    ref.column_dimensions['L'].width = 32

    for i, (r, label) in enumerate(GPU_ROWS, start=2):
        bg = 'FFE0F7FA' if i % 2 == 0 else 'FFFFFFFF'
        ref_cell(r,     i, 11, bold=True, bg=bg)
        ref_cell(label, i, 12, bg=bg)

    ref.row_dimensions[1].height = 22

    return wb

# ─── main ────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='สร้าง picks.xlsx template')
    parser.add_argument('--output', default=str(DEFAULT_OUT),
                        help='path output (default: data/affiliate/picks.xlsx)')
    args = parser.parse_args()

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook()
    wb.save(out)
    print(f'✅ สร้าง {out.relative_to(REPO_ROOT)} แล้ว')
    print(f'   เปิดด้วย: open "{out}"')

if __name__ == '__main__':
    main()
